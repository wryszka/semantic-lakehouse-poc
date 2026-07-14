#!/usr/bin/env python3
"""Build the Excel demo pack: BAU report snapshot + live-connection wiring instructions.

Run: uv run --with openpyxl --native-tls build_excel_pack.py
"""
import json
import os
import subprocess
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HOST = os.environ.get("DATABRICKS_HOST", "https://REPLACE-ME.cloud.databricks.com")
WAREHOUSE = os.environ.get("DATABRICKS_WAREHOUSE_ID", "REPLACE-ME")
HTTP_PATH = f"/sql/1.0/warehouses/{WAREHOUSE}"
BAU_SQL = "SELECT * FROM lr_dev_aws_us_catalog.semantic_lakehouse.vw_bau_premiums_by_sector"
PIVOT_SQL = ("SELECT month, year, sector, gross_written_premium, earned_premium, claims_incurred, "
             "policy_count, claim_count, loss_ratio_at_this_grain, gwp_ytd, gwp_rolling_12m, "
             "loss_ratio_rolling_12m FROM lr_dev_aws_us_catalog.semantic_lakehouse.vw_underwriting_monthly "
             "WHERE year >= 2024")

tok = json.loads(subprocess.run(["databricks", "auth", "token", "-p", os.environ.get("DATABRICKS_PROFILE", "REPLACE-ME")],
                                capture_output=True, text=True).stdout)["access_token"]


def query(stmt):
    body = json.dumps({"warehouse_id": WAREHOUSE, "statement": stmt,
                       "wait_timeout": "50s", "format": "JSON_ARRAY"}).encode()
    req = urllib.request.Request(f"{HOST}/api/2.0/sql/statements/", data=body,
                                 headers={"Authorization": f"Bearer {tok}",
                                          "Content-Type": "application/json"})
    r = json.load(urllib.request.urlopen(req))
    assert r["status"]["state"] == "SUCCEEDED", r["status"]
    cols = [c["name"] for c in r["manifest"]["schema"]["columns"]]
    types = [c["type_name"] for c in r["manifest"]["schema"]["columns"]]
    return cols, types, r["result"].get("data_array", [])


def write_sheet(ws, cols, types, rows):
    hdr_fill = PatternFill("solid", fgColor="1B3139")
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
    for i, row in enumerate(rows, 2):
        for j, (v, t) in enumerate(zip(row, types), 1):
            if v is None:
                continue
            if t in ("DECIMAL", "DOUBLE", "FLOAT"):
                v = float(v)
            elif t in ("INT", "BIGINT", "LONG", "SMALLINT"):
                v = int(v)
            ws.cell(row=i, column=j, value=v)
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(c) + 2)
    ws.freeze_panes = "A2"


wb = Workbook()

# Sheet 1: the BAU report (what the business opens every morning)
ws = wb.active
ws.title = "BAU Premiums by Sector"
cols, types, rows = query(BAU_SQL)
write_sheet(ws, cols, types, rows)

# Sheet 2: pivot-ready extract from the wrapper view
ws2 = wb.create_sheet("Pivot Source (2024+)")
cols, types, rows = query(PIVOT_SQL)
write_sheet(ws2, cols, types, rows)

# Sheet 3: how to wire the live refresh
ws3 = wb.create_sheet("Live Connection Setup")
lines = [
    ("How this workbook goes live (one-time setup, ~2 minutes)", True),
    ("", False),
    ("This snapshot was produced from the governed wrapper view. To make Data > Refresh All", False),
    ("pull live data, add a Power Query connection:", False),
    ("", False),
    ("1. Data > Get Data > From Other Sources > From ODBC (or 'From Databricks' if available)", True),
    (f"   Server hostname:  {HOST.replace('https://', '')}", False),
    (f"   HTTP path:        {HTTP_PATH}", False),
    ("   Authentication:   OAuth (Azure AD / U2M) or personal access token", False),
    ("", False),
    ("2. Paste this SQL as the query:", True),
    (f"   {BAU_SQL}", False),
    ("", False),
    ("3. Load to the 'BAU Premiums by Sector' sheet. Done - the workbook now refreshes", False),
    ("   from the same governed metric definitions as Genie, dashboards and Power BI.", False),
    ("", False),
    ("Power Query (M) equivalent:", True),
    (f'   Odbc.Query("DSN=Databricks", "{BAU_SQL}")', False),
    ("", False),
    ("IMPORTANT - the loss_ratio_* columns are valid only at the row grain.", True),
    ("Never SUM a ratio column in a pivot: recompute as SUM(claims_incurred) / SUM(earned_premium).", False),
    ("The additive columns (premiums, counts) are safe to total.", False),
    ("", False),
    ("OPTION 2 - Databricks Excel Add-in: see the 'Excel Add-in (Preview)' sheet for install and usage.", True),
    ("", False),
    ("About this demo: synthetic Bricksurance SE data, generated for demonstration purposes.", False),
]
for i, (txt, bold) in enumerate(lines, 1):
    c = ws3.cell(row=i, column=1, value=txt)
    if bold:
        c.font = Font(bold=True)
ws3.column_dimensions["A"].width = 110


# Sheet 4: the Databricks Excel Add-in - install and use
ws4 = wb.create_sheet("Excel Add-in (Preview)")
addin_lines = [
    ("Databricks Excel Add-in - Public Preview since 2 March 2026", True),
    ("Reads Unity Catalog tables AND Metric Views directly: browse the catalog, import, pivot,", False),
    ("run optional SQL, and even write data back to Unity Catalog. No ODBC driver, no tokens - SSO sign-in.", False),
    ("Supported: Excel on the web, Excel on Windows (Microsoft 365), Excel on macOS (2019+).", False),
    ("", False),
    ("BEFORE YOU START (once per workspace)", True),
    ("A workspace admin must enable the Excel Connector preview; users need Databricks SQL access,", False),
    ("CAN USE on a SQL warehouse, and SELECT on the data. Manifest/add-in file: see the setup doc below.", False),
    ("", False),
    ("INSTALL - Excel on the web", True),
    ("  1. Open a workbook  >  Home tab  >  Add-ins  >  Advanced  >  Upload My Add-in", False),
    ("  2. Upload the Databricks add-in file, then open the Databricks Add-in from the Add-ins menu", False),
    ("  3. Sign in to Databricks (allow pop-ups); pick the workspace if more than one is configured", False),
    ("", False),
    ("INSTALL - Excel on Windows (desktop)", True),
    ("  1. Create a folder, e.g. C:\\Manifest, and copy the add-in file into it", False),
    ("  2. Share the folder (Properties > Sharing, read/write)", False),
    ("  3. Excel: File > Options > Trust Center > Trust Center Settings > Trusted Add-in Catalogs", False),
    ("     - add the folder path (\\\\YourComputerName\\Manifest), tick 'Show in Menu', restart Excel", False),
    ("  4. Search 'Insert as Add-in' in the title bar, select the Databricks connector, click Add", False),
    ("  5. Open the add-in and sign in", False),
    ("", False),
    ("INSTALL - Excel on macOS (desktop)", True),
    ("  1. Copy the add-in file to ~/Library/Containers/com.microsoft.Excel/Data/Documents/wef", False),
    ("  2. Restart Excel  >  Add-ins  >  My Add-ins  >  Databricks Add-in  >  sign in", False),
    ("", False),
    ("CONNECT AND USE (all platforms)", True),
    ("  1. Home tab  >  Databricks Add-in  >  enter the workspace URL  >  Sign in (SSO)", False),
    ("  2. Browse the catalog and import a table or METRIC VIEW - for this demo:", False),
    ("     lr_dev_aws_us_catalog.semantic_lakehouse.mv_underwriting", False),
    ("  3. Pivot on the imported data, or run a custom SQL query if you prefer", False),
    ("  4. Data refreshes on demand from the same governed definitions as Genie and dashboards", False),
    ("", False),
    ("CENTRAL DEPLOYMENT (IT - for users who cannot install add-ins)", True),
    ("  1. Deploy via the Microsoft 365 admin center (Marketplace app or custom manifest)", False),
    ("  2. Microsoft Entra: grant admin consent to the Databricks enterprise application", False),
    ("     (application ID aaec40b0-c0ae-4211-a98b-6fc160abb71b) if third-party consent is blocked", False),
    ("  3. Allowlist the Databricks workspace URL on firewalls/proxies", False),
    ("", False),
    ("DOCS", True),
    ("  Overview:   https://docs.databricks.com/aws/en/integrations/excel", False),
    ("  Setup:      https://docs.databricks.com/aws/en/integrations/excel-setup", False),
    ("  Import/query: https://docs.databricks.com/aws/en/integrations/excel-query", False),
    ("  Write-back: https://docs.databricks.com/aws/en/integrations/excel-write-back", False),
]
for i, (txt, bold) in enumerate(addin_lines, 1):
    cell = ws4.cell(row=i, column=1, value=txt)
    if bold:
        cell.font = Font(bold=True)
ws4.column_dimensions["A"].width = 110

wb.save("underwriting_bau_report.xlsx")
print("written underwriting_bau_report.xlsx")
